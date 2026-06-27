from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class RawCell:
    row: int
    col: int
    address: str
    value: Any
    display_value: Any | None = None
    fill_color: str | None = None
    bold: bool = False
    merged_range: str | None = None


@dataclass(frozen=True)
class RawMergedRange:
    address: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int


@dataclass(frozen=True)
class RawCellGrid:
    sheet_name: str
    max_row: int
    max_col: int
    cells: list[RawCell]
    merged_ranges: list[RawMergedRange]

    def cell_at(self, row: int, col: int) -> RawCell:
        for cell in self.cells:
            if cell.row == row and cell.col == col:
                return cell
        raise KeyError(f"Cell not found at row={row}, col={col}")


class RawCellGridExtractor:
    def extract(self, path: str) -> list[RawCellGrid]:
        file_path = Path(path)
        suffix = file_path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            return self._extract_openpyxl_workbook(file_path)
        if suffix == ".xls":
            return self._extract_xls_workbook(file_path)
        raise ValueError("RawCellGridExtractor 当前只支持 .xlsx、.xlsm 和 .xls")

    def _extract_openpyxl_workbook(self, file_path: Path) -> list[RawCellGrid]:
        workbook = load_workbook(file_path, read_only=False, data_only=True)
        try:
            return [self._extract_sheet(worksheet) for worksheet in workbook.worksheets]
        finally:
            workbook.close()

    def _extract_sheet(self, worksheet) -> RawCellGrid:
        merged_ranges = [
            RawMergedRange(
                address=str(cell_range),
                min_row=cell_range.min_row,
                max_row=cell_range.max_row,
                min_col=cell_range.min_col,
                max_col=cell_range.max_col,
            )
            for cell_range in worksheet.merged_cells.ranges
        ]
        cells: list[RawCell] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None and not self._cell_has_style_signal(cell):
                    continue
                cells.append(
                    RawCell(
                        row=cell.row,
                        col=cell.column,
                        address=cell.coordinate,
                        value=cell.value,
                        display_value=self._display_value(cell.value, cell.number_format),
                        fill_color=self._fill_color(cell),
                        bold=bool(cell.font and cell.font.bold),
                        merged_range=self._merged_range_for_cell(cell.row, cell.column, merged_ranges),
                    )
                )
        return RawCellGrid(
            sheet_name=worksheet.title,
            max_row=worksheet.max_row,
            max_col=worksheet.max_column,
            cells=cells,
            merged_ranges=merged_ranges,
        )

    def _cell_has_style_signal(self, cell) -> bool:
        return bool(cell.fill and cell.fill.fill_type) or bool(cell.font and cell.font.bold)

    def _fill_color(self, cell) -> str | None:
        if not cell.fill or not cell.fill.fill_type:
            return None
        color = cell.fill.fgColor
        return color.rgb if color and color.type == "rgb" else None

    def _merged_range_for_cell(
        self,
        row: int,
        col: int,
        merged_ranges: list[RawMergedRange],
    ) -> str | None:
        for merged_range in merged_ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
            ):
                return merged_range.address
        return None

    def _extract_xls_workbook(self, file_path: Path) -> list[RawCellGrid]:
        import xlrd

        workbook = xlrd.open_workbook(file_path, formatting_info=True)
        return [
            self._extract_xls_sheet(workbook.sheet_by_index(index), workbook)
            for index in range(workbook.nsheets)
        ]

    def _extract_xls_sheet(self, worksheet, workbook) -> RawCellGrid:
        merged_ranges = [
            RawMergedRange(
                address=self._range_address(row_start + 1, row_end, col_start + 1, col_end),
                min_row=row_start + 1,
                max_row=row_end,
                min_col=col_start + 1,
                max_col=col_end,
            )
            for row_start, row_end, col_start, col_end in getattr(worksheet, "merged_cells", [])
        ]
        cells: list[RawCell] = []
        for row_index in range(worksheet.nrows):
            for col_index in range(worksheet.ncols):
                value = worksheet.cell_value(row_index, col_index)
                if value == "":
                    continue
                row = row_index + 1
                col = col_index + 1
                normalized_value = self._normalize_xls_value(value)
                cells.append(
                    RawCell(
                        row=row,
                        col=col,
                        address=self._address(row, col),
                        value=normalized_value,
                        display_value=self._xls_display_value(
                            normalized_value,
                            worksheet,
                            workbook,
                            row_index,
                            col_index,
                        ),
                        merged_range=self._merged_range_for_cell(row, col, merged_ranges),
                    )
                )
        return RawCellGrid(
            sheet_name=worksheet.name,
            max_row=max(worksheet.nrows, 1),
            max_col=max(worksheet.ncols, 1),
            cells=cells,
            merged_ranges=merged_ranges,
        )

    def _normalize_xls_value(self, value):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    def _xls_display_value(self, value, worksheet, workbook, row_index: int, col_index: int):
        number_format = self._xls_number_format(worksheet, workbook, row_index, col_index)
        return self._display_value(value, number_format)

    def _xls_number_format(self, worksheet, workbook, row_index: int, col_index: int) -> str | None:
        try:
            xf_index = worksheet.cell_xf_index(row_index, col_index)
            xf = workbook.xf_list[xf_index]
            cell_format = workbook.format_map.get(xf.format_key)
        except (AttributeError, IndexError, KeyError, TypeError):
            return None
        return getattr(cell_format, "format_str", None)

    def _display_value(self, value, number_format: str | None):
        if value is None:
            return None
        if not isinstance(value, (int, float)) or isinstance(value, bool):
            return value
        decimals = self._display_decimal_places(number_format or "")
        if decimals is None:
            return value
        return f"{value:.{decimals}f}"

    def _display_decimal_places(self, number_format: str) -> int | None:
        if not number_format or number_format == "General":
            return None
        first_section = number_format.split(";", 1)[0]
        first_section = re.sub(r'"[^"]*"', "", first_section)
        if "." not in first_section:
            return None
        decimal_part = first_section.split(".", 1)[1]
        match = re.match(r"([0#?]+)", decimal_part)
        if not match:
            return None
        return len(match.group(1))

    def _range_address(self, min_row: int, max_row: int, min_col: int, max_col: int) -> str:
        return f"{self._address(min_row, min_col)}:{self._address(max_row, max_col)}"

    def _address(self, row: int, col: int) -> str:
        return f"{self._column_letter(col)}{row}"

    def _column_letter(self, col: int) -> str:
        letters = ""
        while col:
            col, remainder = divmod(col - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters
