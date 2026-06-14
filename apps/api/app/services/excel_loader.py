from dataclasses import dataclass
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook

from app.services.excel_parse_quality import ExcelParseQualityEvaluator
from app.services.excel_structure_detector import (
    ExcelSheetSignals,
    ExcelStructure,
    ExcelStructureDetector,
    MergedCellRange,
)


@dataclass(frozen=True)
class LoadedSheet:
    name: str
    dataframe: pd.DataFrame
    title: str | None = None
    subtitle: str | None = None
    unit: str | None = None


class ExcelLoader:
    """Excel/CSV 解析服务。这里只做结构化解析，不调用 AI。"""

    def __init__(self) -> None:
        self.structure_detector = ExcelStructureDetector()
        self.quality_evaluator = ExcelParseQualityEvaluator()

    def load(self, path: str) -> list[LoadedSheet]:
        file_path = Path(path)
        extension = file_path.suffix.lower()

        if extension == ".csv":
            dataframe = pd.read_csv(file_path)
            return [LoadedSheet(name=file_path.stem, dataframe=self._normalize_dataframe(dataframe))]

        if extension in {".xlsx", ".xlsm"}:
            return self._load_excel_workbook(file_path, engine="openpyxl")

        if extension == ".xls":
            return self._load_excel_workbook(file_path, engine="xlrd")

        if extension == ".xlsb":
            return self._load_excel_workbook(file_path, engine="pyxlsb")

        if extension == ".ods":
            return self._load_excel_workbook(file_path, engine="odf")

        if extension == ".et":
            return self._load_wps_et(file_path)

        raise ValueError("只支持 .xlsx、.xls、.xlsm、.xlsb、.et、.ods 和 .csv 文件")

    def _load_excel_workbook(self, file_path: Path, engine: str) -> list[LoadedSheet]:
        sheets = pd.read_excel(file_path, sheet_name=None, engine=engine, header=None)
        signals_by_sheet = (
            self._load_openpyxl_signals(file_path)
            if engine == "openpyxl"
            else {}
        )
        loaded_sheets: list[LoadedSheet] = []
        for name, dataframe in sheets.items():
            loaded_sheets.extend(
                sheet
                for sheet in self._normalize_sheet_regions(
                    name,
                    dataframe,
                    signals=signals_by_sheet.get(name),
                )
                if self._is_importable_dataframe(sheet.dataframe)
            )
        return loaded_sheets

    def _load_wps_et(self, file_path: Path) -> list[LoadedSheet]:
        errors: list[str] = []
        for engine in ("openpyxl", "xlrd"):
            try:
                return self._load_excel_workbook(file_path, engine=engine)
            except Exception as exc:
                errors.append(f"{engine}: {exc}")

        details = "；".join(errors)
        raise ValueError(f".et 文件解析失败，请另存为 .xlsx 或 .xls 后重试。{details}")

    def _normalize_dataframe(
        self,
        dataframe: pd.DataFrame,
        signals: ExcelSheetSignals | None = None,
    ) -> pd.DataFrame:
        normalized = dataframe.dropna(axis=1, how="all").copy()
        normalized, _ = self._promote_embedded_header(normalized, signals=signals)
        normalized = self._normalize_text_values(normalized)
        normalized.columns = self._normalize_columns([str(column) for column in normalized.columns])
        return normalized

    def _normalize_sheet(
        self,
        name: str,
        dataframe: pd.DataFrame,
        signals: ExcelSheetSignals | None = None,
    ) -> LoadedSheet:
        normalized = dataframe.dropna(axis=1, how="all").copy()
        metadata = self._extract_sheet_metadata(normalized, signals=signals)
        normalized, _ = self._promote_embedded_header(normalized, signals=signals)
        normalized = self._normalize_text_values(normalized)
        normalized.columns = self._normalize_columns([str(column) for column in normalized.columns])
        return LoadedSheet(name=name, dataframe=normalized, **metadata)

    def _normalize_sheet_regions(
        self,
        name: str,
        dataframe: pd.DataFrame,
        signals: ExcelSheetSignals | None = None,
    ) -> list[LoadedSheet]:
        normalized = dataframe.dropna(axis=1, how="all").copy()
        regions = self.structure_detector.detect_regions(normalized, signals=signals)
        regions = self._merge_header_only_year_regions(normalized, regions)
        if not regions:
            loaded_sheet = self._normalize_sheet(name, dataframe, signals=signals)
            return [loaded_sheet] if self._is_importable_dataframe(loaded_sheet.dataframe) else []

        loaded_sheets = []
        for index, region in enumerate(regions, start=1):
            header_rows = [
                [self._cell_to_text(value) for value in normalized.iloc[row_index]]
                for row_index in range(region.header_start or 0, (region.header_end or 0) + 1)
            ]
            data_start = self._skip_blank_rows(normalized, region.data_start)
            data_end = self._find_region_data_end(normalized, data_start)
            region_frame = normalized.iloc[data_start : data_end].copy()
            region_frame.columns = self._compose_header_rows(header_rows)
            region_frame = region_frame.dropna(how="all").dropna(axis=1, how="all")
            region_frame = self._normalize_text_values(region_frame)
            region_frame.columns = self._normalize_columns(
                [str(column) for column in region_frame.columns]
            )
            if not self._is_importable_dataframe(region_frame):
                continue
            if not self.quality_evaluator.evaluate(region_frame).is_importable:
                continue
            metadata = self._extract_region_metadata(normalized, region)
            loaded_sheets.append(
                LoadedSheet(
                    name=name if index == 1 else f"{name} #{index}",
                    dataframe=region_frame,
                    **metadata,
                )
            )
        return loaded_sheets

    def _merge_header_only_year_regions(
        self,
        dataframe: pd.DataFrame,
        regions: list[ExcelStructure],
    ) -> list[ExcelStructure]:
        merged: list[ExcelStructure] = []
        index = 0
        while index < len(regions):
            region = regions[index]
            next_region = regions[index + 1] if index + 1 < len(regions) else None
            if (
                next_region
                and "year_header_cells" in region.reasons
                and not self._region_has_data(dataframe, region)
                and self._region_has_data(dataframe, next_region)
            ):
                merged.append(
                    ExcelStructure(
                        header_start=region.header_start,
                        header_end=region.header_end,
                        data_start=self._merged_region_data_start(next_region),
                        title_rows=region.title_rows,
                        confidence=region.confidence,
                        reasons=[*region.reasons, "merged_print_spacer_region"],
                    )
                )
                index += 2
                continue
            merged.append(region)
            index += 1
        return merged

    def _merged_region_data_start(self, next_region: ExcelStructure) -> int:
        starts = [
            row
            for row in [*next_region.title_rows, next_region.header_start, next_region.data_start]
            if row is not None
        ]
        return min(starts) if starts else next_region.data_start

    def _region_has_data(self, dataframe: pd.DataFrame, region: ExcelStructure) -> bool:
        data_end = self._find_region_data_end(dataframe, region.data_start)
        if region.data_start >= data_end:
            return False
        frame = dataframe.iloc[region.data_start:data_end].dropna(how="all").dropna(axis=1, how="all")
        return self._is_importable_dataframe(frame)

    def _extract_region_metadata(
        self,
        dataframe: pd.DataFrame,
        region,
    ) -> dict[str, str | None]:
        title = None
        subtitle = None
        unit = None
        for row_index in self._metadata_row_indexes(dataframe, region):
            values = [self._cell_to_text(value) for value in dataframe.iloc[row_index]]
            texts = [value for value in values if value]
            if not texts:
                continue
            joined = " ".join(texts)
            if unit is None:
                unit = self._extract_unit_text(texts)
            if subtitle is None:
                subtitle_match = re.search(r"[（(]\s*\d{4}\s*年\s*[）)]", joined)
                if subtitle_match:
                    subtitle = subtitle_match.group(0)
            if title is None:
                title_candidates = [
                    text
                    for text in texts
                    if self._looks_like_title_text(text)
                ]
                if title_candidates:
                    title = max(title_candidates, key=len)
        return {"title": title, "subtitle": subtitle, "unit": unit}

    def _metadata_row_indexes(self, dataframe: pd.DataFrame, region: ExcelStructure) -> list[int]:
        header_start = region.header_start if region.header_start is not None else region.data_start
        start = max(0, header_start - 8)
        indexes = set(region.title_rows)
        for row_index in range(header_start - 1, start - 1, -1):
            values = list(dataframe.iloc[row_index])
            texts = [self._cell_to_text(value) for value in values if self._has_value(value)]
            if not texts:
                continue
            if self._row_has_non_year_numeric(values):
                break
            if len(texts) <= 2 or self._extract_unit_text(texts):
                indexes.add(row_index)
        return sorted(indexes)

    def _promote_embedded_header(
        self,
        dataframe: pd.DataFrame,
        signals: ExcelSheetSignals | None = None,
    ) -> tuple[pd.DataFrame, int | None]:
        regions = self.structure_detector.detect_regions(dataframe, signals=signals)
        structure = regions[0] if regions else self.structure_detector.detect(dataframe, signals=signals)
        if structure.header_start is None or structure.header_end is None:
            return dataframe, None

        header_rows = [
            [self._cell_to_text(value) for value in dataframe.iloc[index]]
            for index in range(structure.header_start, structure.header_end + 1)
        ]
        data_start = self._skip_blank_rows(dataframe, structure.data_start)
        data_end = self._find_region_data_end(dataframe, data_start)
        promoted = dataframe.iloc[data_start : data_end].copy()
        promoted.columns = self._compose_header_rows(header_rows)
        return promoted.dropna(how="all").dropna(axis=1, how="all"), structure.header_start

    def _find_region_data_end(self, dataframe: pd.DataFrame, data_start: int) -> int:
        for index in range(data_start, len(dataframe.index)):
            if not any(self._has_value(value) for value in list(dataframe.iloc[index])):
                return index
        return len(dataframe.index)

    def _skip_blank_rows(self, dataframe: pd.DataFrame, row_index: int) -> int:
        while row_index < len(dataframe.index):
            if any(self._has_value(value) for value in list(dataframe.iloc[row_index])):
                return row_index
            row_index += 1
        return row_index

    def _compose_header_rows(self, header_rows: list[list[str]]) -> list[str]:
        if len(header_rows) == 1:
            return header_rows[0]

        header_rows = self._merge_wrapped_header_rows(header_rows)
        inherited_rows = [self._fill_header_groups(row) for row in header_rows]
        headers = []
        for column_index in range(len(inherited_rows[0])):
            parts = []
            for row in inherited_rows:
                value = row[column_index]
                if value and value not in parts:
                    parts.append(value)
            headers.append("_".join(parts))
        return headers

    def _merge_wrapped_header_rows(self, header_rows: list[list[str]]) -> list[list[str]]:
        merged: list[list[str]] = []
        for row in header_rows:
            non_empty_indexes = [index for index, value in enumerate(row) if value]
            if (
                len(non_empty_indexes) == 1
                and merged
                and merged[-1][non_empty_indexes[0]]
            ):
                column_index = non_empty_indexes[0]
                merged[-1][column_index] += row[column_index]
                continue
            merged.append(row.copy())
        return merged

    def _fill_header_groups(self, row: list[str]) -> list[str]:
        filled = []
        current = ""
        for value in row:
            if value:
                current = value
            filled.append(value or current)
        return filled

    def _extract_sheet_metadata(
        self,
        dataframe: pd.DataFrame,
        signals: ExcelSheetSignals | None = None,
    ) -> dict[str, str | None]:
        header_index = self._find_header_index(dataframe, signals=signals)
        search_end = header_index if header_index is not None else min(6, len(dataframe.index))
        title = None
        subtitle = None
        unit = None

        for index in range(search_end):
            values = [self._cell_to_text(value) for value in dataframe.iloc[index]]
            texts = [value for value in values if value]
            if not texts:
                continue

            joined = " ".join(texts)
            if unit is None:
                unit = self._extract_unit_text(texts)

            if subtitle is None:
                subtitle_match = re.search(r"[（(]\s*\d{4}\s*年\s*[）)]", joined)
                if subtitle_match:
                    subtitle = subtitle_match.group(0)

            if title is None:
                title_candidates = [
                    text
                    for text in texts
                    if self._looks_like_title_text(text)
                ]
                if title_candidates:
                    title = max(title_candidates, key=len)

        return {"title": title, "subtitle": subtitle, "unit": unit}

    def _find_header_index(
        self,
        dataframe: pd.DataFrame,
        signals: ExcelSheetSignals | None = None,
    ) -> int | None:
        regions = self.structure_detector.detect_regions(dataframe, signals=signals)
        if regions:
            return regions[0].header_start
        return self.structure_detector.detect(dataframe, signals=signals).header_start

    def _load_openpyxl_signals(self, file_path: Path) -> dict[str, ExcelSheetSignals]:
        workbook = load_workbook(file_path, read_only=False, data_only=True)
        signals_by_sheet: dict[str, ExcelSheetSignals] = {}
        for worksheet in workbook.worksheets:
            merged_ranges = [
                MergedCellRange(
                    start_row=cell_range.min_row - 1,
                    end_row=cell_range.max_row - 1,
                    start_col=cell_range.min_col - 1,
                    end_col=cell_range.max_col - 1,
                )
                for cell_range in worksheet.merged_cells.ranges
            ]
            styled_rows = {
                row_index - 1
                for row_index in range(1, worksheet.max_row + 1)
                if self._row_has_header_style(worksheet[row_index])
            }
            signals_by_sheet[worksheet.title] = ExcelSheetSignals(
                merged_ranges=merged_ranges,
                styled_rows=styled_rows,
            )
        workbook.close()
        return signals_by_sheet

    def _row_has_header_style(self, cells) -> bool:
        styled_count = 0
        non_empty_count = 0
        for cell in cells:
            if not self._has_value(cell.value):
                continue
            non_empty_count += 1
            has_fill = bool(cell.fill and cell.fill.fill_type)
            has_border = bool(
                cell.border
                and any(
                    side.style
                    for side in (
                        cell.border.left,
                        cell.border.right,
                        cell.border.top,
                        cell.border.bottom,
                    )
                )
            )
            if cell.font.bold or has_fill or has_border:
                styled_count += 1
        return non_empty_count > 0 and styled_count >= max(1, non_empty_count // 2)

    def _normalize_text_values(self, dataframe: pd.DataFrame) -> pd.DataFrame:
        normalized = dataframe.copy()
        for column in normalized.columns:
            normalized[column] = normalized[column].map(self._normalize_cell_value)
        return normalized

    def _normalize_columns(self, columns: list[str]) -> list[str]:
        seen: dict[str, int] = {}
        result: list[str] = []

        for index, column in enumerate(columns, start=1):
            column = self._normalize_cell_text(column)
            base = re.sub(r"\W+", "_", column.strip().lower()).strip("_")
            base = re.sub(r"_+", "_", base)
            if not base or base.startswith("unnamed"):
                base = f"column_{index}"

            count = seen.get(base, 0)
            seen[base] = count + 1
            result.append(base if count == 0 else f"{base}_{count + 1}")

        return result

    def _normalize_cell_value(self, value):
        if isinstance(value, str):
            return self._normalize_cell_text(value)
        return value

    def _normalize_cell_text(self, value: str) -> str:
        text = value.replace("\u3000", " ")
        text = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])", "", text)
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    def _has_value(self, value) -> bool:
        if pd.isna(value):
            return False
        return bool(str(value).strip())

    def _cell_to_text(self, value) -> str:
        if pd.isna(value):
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return self._normalize_cell_text(str(value))

    def _looks_numeric(self, value: str) -> bool:
        return bool(re.fullmatch(r"[-+]?\d+(\.\d+)?%?", value.strip()))

    def _row_has_non_year_numeric(self, values: list) -> bool:
        for value in values:
            if not self._has_value(value):
                continue
            text = self._cell_to_text(value)
            if self._looks_numeric(text) and not re.fullmatch(r"(?:19|20)\d{2}(?:\.0)?", text):
                return True
        return False

    def _extract_unit_text(self, texts: list[str]) -> str | None:
        for text in texts:
            stripped = text.strip()
            if ":" in stripped or "：" in stripped:
                candidate = re.split(r"[:：]", stripped, maxsplit=1)[1].strip().strip("()（） ")
                if self._looks_like_delimited_unit_token(candidate):
                    return candidate
            inner = stripped.strip("()（） ")
            if self._looks_like_symbol_unit_token(inner):
                return inner
        return None

    def _looks_like_delimited_unit_token(self, value: str) -> bool:
        if not value:
            return False
        if len(value) > 8:
            return False
        if re.search(r"[A-Za-z]", value):
            return False
        return bool(re.search(r"[^0-9\s]", value))

    def _looks_like_symbol_unit_token(self, value: str) -> bool:
        if not value:
            return False
        if len(value) > 8:
            return False
        return not re.search(r"[\w\u4e00-\u9fff]", value)

    def _looks_like_title_text(self, text: str) -> bool:
        if self._extract_unit_text([text]):
            return False
        if re.fullmatch(r"[（(]?\s*\d{4}\s*年\s*[）)]?", text):
            return False
        if re.fullmatch(r"\(?\s*[^A-Za-z0-9\u4e00-\u9fff]{1,6}\s*\)?", text):
            return False
        return True

    def _is_importable_dataframe(self, dataframe: pd.DataFrame) -> bool:
        return (
            len(dataframe.index) > 0
            and len(dataframe.columns) > 0
            and self.quality_evaluator.evaluate(dataframe).is_importable
        )
