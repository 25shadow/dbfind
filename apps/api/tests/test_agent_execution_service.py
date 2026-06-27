from datetime import UTC, datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.repositories.file_repository import FileRepository
from app.repositories.sheet_repository import SheetRepository
from app.schemas.agent import AgentPlan, AgentPreview, AgentStep
from app.schemas.query import QueryResponse
from app.services.agent_execution_service import AgentExecutionService
from app.services.duckdb_service import DuckdbService


def test_execute_selected_file_plan_generates_styled_workbook(temp_workspace):
    file_id = "file-1"
    original_path = temp_workspace / "files" / "source.xlsx"
    original_path.parent.mkdir(parents=True)
    original_path.write_bytes(b"original")

    FileRepository().create_file(
        file_id=file_id,
        name="source.xlsx",
        path=str(original_path),
        file_hash="hash",
        status="ready",
        created_at=datetime.now(UTC).isoformat(),
    )
    SheetRepository().replace_sheets(
        file_id,
        [
            {
                "id": f"{file_id}_1",
                "name": "销售明细",
                "table_name": "sales",
                "row_count": 2,
                "column_count": 2,
                "title": None,
                "subtitle": None,
                "unit": None,
            }
        ],
    )
    DuckdbService().write_dataframe(
        DuckdbService().database_path_for_file(file_id),
        "sales",
        pd.DataFrame(
            [
                {"客户": "A", "金额": 12.345},
                {"客户": "B", "金额": 67.891},
            ]
        ),
    )
    plan = AgentPlan(
        intent="excel_operation",
        scope="selected",
        summary="生成新工作簿",
        requiresConfirmation=True,
        riskLevel="medium",
        steps=[
            AgentStep(tool="workbook_writer", purpose="生成工作簿", params={}),
            AgentStep(tool="workbook_style", purpose="设计表格", params={}),
        ],
        preview=AgentPreview(),
    )

    result = AgentExecutionService().execute(plan=plan, file_id=file_id)

    output_path = temp_workspace / "generated" / result.output_id
    assert result.status == "completed"
    assert result.file_name.endswith(".xlsx")
    assert result.download_url == f"/api/agent/generated/{result.output_id}"
    assert output_path.exists()
    assert original_path.read_bytes() == b"original"

    workbook = load_workbook(output_path)
    sheet = workbook["销售明细"]
    assert sheet.freeze_panes == "A2"
    assert sheet["A1"].font.bold is True
    assert sheet["A2"].value == "A"
    assert sheet["B3"].value == 67.891


def test_execute_selected_file_plan_applies_transform_and_style_params(temp_workspace):
    file_id = "file-2"
    original_path = temp_workspace / "files" / "source.xlsx"
    original_path.parent.mkdir(parents=True)
    original_path.write_bytes(b"original")

    FileRepository().create_file(
        file_id=file_id,
        name="source.xlsx",
        path=str(original_path),
        file_hash="hash-2",
        status="ready",
        created_at=datetime.now(UTC).isoformat(),
    )
    SheetRepository().replace_sheets(
        file_id,
        [
            {
                "id": f"{file_id}_1",
                "name": "行政区划",
                "table_name": "districts",
                "row_count": 3,
                "column_count": 3,
                "title": None,
                "subtitle": None,
                "unit": None,
            }
        ],
    )
    DuckdbService().write_dataframe(
        DuckdbService().database_path_for_file(file_id),
        "districts",
        pd.DataFrame(
            [
                {"市别": "云浮", "街道": 8.123, "备注": "保留"},
                {"市别": "韶关", "街道": 10.987, "备注": "保留"},
                {"市别": "广州", "街道": 11.111, "备注": "排除"},
            ]
        ),
    )
    plan = AgentPlan(
        intent="excel_operation",
        scope="selected",
        summary="生成韶关和云浮街道数对比表",
        requiresConfirmation=True,
        riskLevel="medium",
        steps=[
            AgentStep(
                tool="dataframe_transform",
                purpose="筛选并格式化结果",
                params={
                    "operations": [
                        {"type": "filter_in", "column": "市别", "values": ["韶关", "云浮"]},
                        {"type": "select_columns", "columns": ["市别", "街道"]},
                        {"type": "sort_values", "by": ["街道"], "ascending": False},
                        {"type": "round", "columns": ["街道"], "decimals": 2},
                    ]
                },
            ),
            AgentStep(
                tool="workbook_writer",
                purpose="写入新工作簿",
                params={"sheetName": "街道数对比"},
            ),
            AgentStep(
                tool="workbook_style",
                purpose="设计表格",
                params={
                    "numberFormats": {"街道": "0.00"},
                    "charts": [
                        {
                            "type": "column",
                            "title": "街道数对比",
                            "categoriesColumn": "市别",
                            "valuesColumn": "街道",
                            "position": "D2",
                        }
                    ],
                },
            ),
        ],
        preview=AgentPreview(),
    )

    result = AgentExecutionService().execute(plan=plan, file_id=file_id)

    output_path = temp_workspace / "generated" / result.output_id
    workbook = load_workbook(output_path)
    sheet = workbook["街道数对比"]
    assert sheet.auto_filter.ref == "A1:B3"
    assert sheet["A2"].value == "韶关"
    assert sheet["B2"].value == 10.99
    assert sheet["B2"].number_format == "0.00"
    assert sheet["A3"].value == "云浮"
    assert sheet["B3"].value == 8.12
    assert sheet.max_column == 2
    assert sheet.max_row == 3
    assert len(sheet._charts) == 1


def test_preview_selected_file_plan_returns_transformed_rows_and_design_summary(temp_workspace):
    file_id = "file-preview"
    original_path = temp_workspace / "files" / "source.xlsx"
    original_path.parent.mkdir(parents=True)
    original_path.write_bytes(b"original")

    FileRepository().create_file(
        file_id=file_id,
        name="source.xlsx",
        path=str(original_path),
        file_hash="hash-preview",
        status="ready",
        created_at=datetime.now(UTC).isoformat(),
    )
    SheetRepository().replace_sheets(
        file_id,
        [
            {
                "id": f"{file_id}_1",
                "name": "销售",
                "table_name": "sales",
                "row_count": 3,
                "column_count": 2,
                "title": None,
                "subtitle": None,
                "unit": None,
            }
        ],
    )
    DuckdbService().write_dataframe(
        DuckdbService().database_path_for_file(file_id),
        "sales",
        pd.DataFrame(
            [
                {"城市": "广州", "金额": 10.123},
                {"城市": "佛山", "金额": 8.456},
                {"城市": "韶关", "金额": 4.111},
            ]
        ),
    )
    plan = AgentPlan(
        intent="excel_operation",
        scope="selected",
        summary="生成销售预览",
        requiresConfirmation=True,
        riskLevel="medium",
        steps=[
            AgentStep(
                tool="dataframe_transform",
                purpose="筛选并保留小数",
                params={
                    "operations": [
                        {"type": "query", "expression": "`金额` >= 8"},
                        {"type": "round", "columns": ["金额"], "decimals": 1},
                    ]
                },
            ),
            AgentStep(tool="workbook_writer", purpose="写入新工作簿", params={"sheetName": "销售预览"}),
            AgentStep(
                tool="workbook_style",
                purpose="样式",
                params={"asTable": True, "numberFormats": {"金额": "0.0"}},
            ),
        ],
        preview=AgentPreview(),
    )

    preview = AgentExecutionService().preview(plan=plan, file_id=file_id)

    assert preview.status == "preview"
    assert preview.sheets[0].sheet_name == "销售预览"
    assert preview.sheets[0].columns == ["城市", "金额"]
    assert preview.sheets[0].rows == [{"城市": "广州", "金额": 10.1}, {"城市": "佛山", "金额": 8.5}]
    assert preview.affected_rows == 2
    assert preview.affected_columns == ["城市", "金额"]
    assert preview.design["asTable"] is True
    assert preview.design["numberFormats"] == {"金额": "0.0"}


def test_query_generated_workbook_includes_source_context(temp_workspace):
    class FakeQueryService:
        def run(self, request, record_history=True):
            return QueryResponse(
                queryId="query-1",
                fileId="",
                scope=request.scope,
                question=request.question,
                sql="select ...",
                columns=["地区", "产值_万元"],
                rows=[
                    {"地区": "武江区", "产值_万元": 29487.89},
                    {"地区": "仁化县", "产值_万元": 103524.27},
                ],
                explanation="",
                createdAt=datetime.now(UTC).isoformat(),
                sources=[
                    {
                        "collectionName": "2024年统计资料",
                        "fileName": "农业产值.xlsx",
                        "sheetName": "Sheet1",
                        "sheetTitle": "季度农业产值",
                    },
                    {
                        "collectionName": "2026年资料",
                        "fileName": "补充数据.xlsx",
                        "sheetName": "Sheet2",
                    },
                ],
            )

    plan = AgentPlan(
        intent="excel_operation",
        scope="all",
        summary="生成对比表",
        requiresConfirmation=True,
        riskLevel="medium",
        steps=[
            AgentStep(
                tool="query",
                purpose="查询数据",
                params={"question": "生成对比表", "scope": "all", "sheetName": "对比表"},
            ),
            AgentStep(tool="workbook_writer", purpose="生成工作簿", params={"sheetName": "对比表"}),
            AgentStep(tool="workbook_style", purpose="设计表格", params={"asTable": True}),
        ],
        preview=AgentPreview(),
    )
    service = AgentExecutionService()
    service.query_service = FakeQueryService()

    preview = service.preview(plan=plan, file_id=None)

    assert preview.sheets[0].columns == ["地区", "产值_万元", "来源"]
    assert preview.sheets[0].rows[0]["来源"] == "2024年统计资料 / 农业产值.xlsx / 季度农业产值"
    assert preview.sheets[0].rows[1]["来源"] == "2026年资料 / 补充数据.xlsx / Sheet2"

    result = service.execute(plan=plan, file_id=None)
    workbook = load_workbook(temp_workspace / "generated" / result.output_id)
    sheet = workbook["对比表"]
    assert sheet["C1"].value == "来源"
    assert sheet["C2"].value == "2024年统计资料 / 农业产值.xlsx / 季度农业产值"
    assert sheet["C3"].value == "2026年资料 / 补充数据.xlsx / Sheet2"


def test_execute_all_files_query_plan_writes_query_result_workbook(temp_workspace, monkeypatch):
    captured = {}

    def fake_run(self, payload, record_history=True):
        captured["payload"] = payload
        captured["record_history"] = record_history
        return QueryResponse(
            queryId="query-1",
            fileId="all",
            scope="all",
            question=payload.question,
            sql=(
                'SELECT city_name, street_count, '
                '143 - 11 AS "差值（广州 - 佛山）" FROM t'
            ),
            columns=["city_name", "street_count", "差值（广州 - 佛山）"],
            rows=[
                {"city_name": "广州", "street_count": 143, "差值（广州 - 佛山）": 132},
                {"city_name": "佛山", "street_count": 11, "差值（广州 - 佛山）": 132},
            ],
            explanation="查询已完成。",
            createdAt=datetime.now(UTC).isoformat(),
            wasRepaired=False,
            sources=[],
        )

    monkeypatch.setattr("app.services.agent_execution_service.QueryService.run", fake_run)
    plan = AgentPlan(
        intent="excel_operation",
        scope="all",
        summary="生成广州和佛山街道数量计算表",
        requiresConfirmation=True,
        riskLevel="medium",
        steps=[
            AgentStep(
                tool="query",
                purpose="查询广州和佛山街道数量并由 SQL 生成差值",
                params={"question": "查询广州和佛山街道数量", "scope": "all"},
            ),
            AgentStep(
                tool="workbook_writer",
                purpose="生成计算表格",
                params={"sheetName": "街道数量计算表"},
            ),
            AgentStep(
                tool="workbook_style",
                purpose="设置数字格式",
                params={
                    "asTable": True,
                    "tableStyle": "Table Style Medium 2",
                    "numberFormats": {"street_count": "0", "差值（广州 - 佛山）": "0"},
                    "conditionalFormats": [
                        {
                            "column": "差值（广州 - 佛山）",
                            "type": "cell",
                            "criteria": ">",
                            "value": 0,
                            "format": {"bg_color": "#E2F0D9"},
                        }
                    ],
                },
            ),
        ],
        preview=AgentPreview(),
    )

    result = AgentExecutionService().execute(plan=plan, file_id=None)

    assert captured["payload"].scope == "all"
    assert captured["payload"].file_id is None
    assert captured["record_history"] is False
    output_path = temp_workspace / "generated" / result.output_id
    workbook = load_workbook(output_path)
    sheet = workbook["街道数量计算表"]
    assert [cell.value for cell in sheet[1]] == ["city_name", "street_count", "差值（广州 - 佛山）"]
    assert sheet["A2"].value == "广州"
    assert sheet["B2"].value == 143
    assert sheet["C2"].value == 132
    assert sheet["A3"].value == "佛山"
    assert sheet["C3"].value == 132
    assert len(sheet.tables) == 1
    assert len(sheet.conditional_formatting) == 1


def test_preview_rejects_dangerous_dataframe_expression_before_loading_file(temp_workspace):
    plan = AgentPlan(
        intent="excel_operation",
        scope="selected",
        summary="危险预览",
        requiresConfirmation=True,
        riskLevel="medium",
        steps=[
            AgentStep(
                tool="dataframe_transform",
                purpose="危险表达式",
                params={
                    "operations": [
                        {
                            "type": "query",
                            "expression": "__import__('os').system('touch /tmp/dbfind-owned')",
                        }
                    ]
                },
            )
        ],
        preview=AgentPreview(),
    )

    with pytest.raises(ValueError, match="表达式包含不允许的内容"):
        AgentExecutionService().preview(plan=plan, file_id="missing-file")


def test_execute_rejects_writer_file_path_before_creating_output(temp_workspace):
    plan = AgentPlan(
        intent="excel_operation",
        scope="selected",
        summary="指定路径",
        requiresConfirmation=True,
        riskLevel="medium",
        steps=[
            AgentStep(
                tool="workbook_writer",
                purpose="写入指定路径",
                params={"outputPath": "/tmp/out.xlsx"},
            )
        ],
        preview=AgentPreview(),
    )

    with pytest.raises(ValueError, match="不能指定文件路径"):
        AgentExecutionService().execute(plan=plan, file_id="missing-file")
