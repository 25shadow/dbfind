from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.services.excel_cell_grid import RawCellGridExtractor


def test_raw_cell_grid_extractor_preserves_values_styles_and_merges(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2-6"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "1-8 各市法人和产业活动单位数"
    sheet["A5"] = "市别"
    sheet["B5"] = "City"
    sheet["C5"] = 2023
    sheet["C6"] = "法人单位数"
    sheet["C7"] = "Corporate"
    sheet["C8"] = "Units"
    sheet["A10"] = "韶关"
    sheet["B10"] = "Shaoguan"
    sheet["C10"] = 50698
    sheet["D10"] = 12.34567
    sheet["D10"].number_format = "0.00"
    sheet["A5"].fill = PatternFill("solid", fgColor="FFFF99")

    path = tmp_path / "grid.xlsx"
    workbook.save(path)

    grids = RawCellGridExtractor().extract(str(path))

    grid = grids[0]
    assert grid.sheet_name == "2-6"
    assert grid.max_row == 10
    assert grid.max_col == 6
    assert grid.merged_ranges[0].address == "A1:F1"
    assert grid.cell_at(5, 1).value == "市别"
    assert grid.cell_at(5, 1).fill_color == "00FFFF99"
    assert grid.cell_at(10, 3).value == 50698
    assert grid.cell_at(10, 4).value == 12.34567
    assert grid.cell_at(10, 4).display_value == "12.35"


def test_raw_cell_grid_extractor_supports_xls_with_xlrd(monkeypatch, tmp_path):
    class FakeFormat:
        format_str = "0.00"

    class FakeXf:
        format_key = 7

    class FakeSheet:
        name = "OldSheet"
        nrows = 3
        ncols = 2
        merged_cells = [(0, 1, 0, 2)]

        def cell_value(self, row, col):
            values = {
                (0, 0): "Title",
                (1, 0): "Name",
                (1, 1): 2024.0,
                (2, 0): "Alpha",
                (2, 1): 12.5,
            }
            return values.get((row, col), "")

        def cell_xf_index(self, row, col):
            return 0 if (row, col) == (2, 1) else None

    class FakeBook:
        nsheets = 1
        xf_list = [FakeXf()]
        format_map = {7: FakeFormat()}

        def sheet_by_index(self, index):
            assert index == 0
            return FakeSheet()

    def fake_open_workbook(path, formatting_info):
        assert str(path).endswith(".xls")
        assert formatting_info is True
        return FakeBook()

    monkeypatch.setattr("xlrd.open_workbook", fake_open_workbook)
    path = tmp_path / "legacy.xls"
    path.write_bytes(b"fake")

    grid = RawCellGridExtractor().extract(str(path))[0]

    assert grid.sheet_name == "OldSheet"
    assert grid.max_row == 3
    assert grid.max_col == 2
    assert grid.merged_ranges[0].address == "A1:B1"
    assert grid.cell_at(1, 1).merged_range == "A1:B1"
    assert grid.cell_at(2, 2).value == 2024
    assert grid.cell_at(3, 2).value == 12.5
    assert grid.cell_at(3, 2).display_value == "12.50"
