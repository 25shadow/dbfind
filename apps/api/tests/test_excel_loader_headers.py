from copy import copy

import pandas as pd
from openpyxl import Workbook

from app.services.excel_loader import ExcelLoader


def test_promotes_grouped_two_row_headers() -> None:
    dataframe = pd.DataFrame(
        [
            ["标题", None, None, None, None, None],
            ["指标", "单位", "2021", "分组表头", None, None],
            [None, None, None, "年份", "数值", "比例"],
            ["甲项目", "万吨", 12.3, 1997, 45.6, 27.0],
        ]
    )

    loaded = ExcelLoader()._normalize_sheet("Sheet1", dataframe)

    assert list(loaded.dataframe.columns) == [
        "指标",
        "单位",
        "2021",
        "分组表头_年份",
        "分组表头_数值",
        "分组表头_比例",
    ]
    assert loaded.dataframe.iloc[0].to_dict()["分组表头_年份"] == 1997


def test_promotes_wrapped_grouped_headers_and_integer_years() -> None:
    dataframe = pd.DataFrame(
        [
            ["标题", None, None, None, None, None],
            ["指标", "单位", 2021.0, "分组表头", None, None],
            [None, None, None, "年份", "数值", "长表头上半段"],
            [None, None, None, None, None, "下半段"],
            ["甲项目", "万吨", 12.3, 1997, 45.6, 27.0],
        ]
    )

    loaded = ExcelLoader()._normalize_sheet("Sheet1", dataframe)

    assert list(loaded.dataframe.columns) == [
        "指标",
        "单位",
        "2021",
        "分组表头_年份",
        "分组表头_数值",
        "分组表头_长表头上半段下半段",
    ]


def test_promotes_statistical_table_header_with_year_columns() -> None:
    dataframe = pd.DataFrame(
        [
            ["1-2 农业主要指标", None, None, None, None, None, None, None, None],
            ["指标", "单位", 2000.0, 2005.0, 2010.0, 2015.0, 2019.0, 2020.0, 2021.0],
            ["乡村户数", "万户", 1419.91, 1540.84, 1686.62, 1689.97, 1723.91, 1751.30, 1835.5],
            ["农业总产值", "亿元", 1701.18, 2447.57, 3697.18, 5303.63, 7175.89, 7901.92, 8305.84],
        ]
    )

    loaded = ExcelLoader()._normalize_sheet("Sheet1", dataframe)

    assert loaded.title == "1-2 农业主要指标"
    assert list(loaded.dataframe.columns) == [
        "指标",
        "单位",
        "2000",
        "2005",
        "2010",
        "2015",
        "2019",
        "2020",
        "2021",
    ]
    assert loaded.dataframe.iloc[0].to_dict() == {
        "指标": "乡村户数",
        "单位": "万户",
        "2000": 1419.91,
        "2005": 1540.84,
        "2010": 1686.62,
        "2015": 1689.97,
        "2019": 1723.91,
        "2020": 1751.30,
        "2021": 1835.5,
    }


def test_loader_uses_first_detected_region_when_sheet_contains_multiple_tables() -> None:
    dataframe = pd.DataFrame(
        [
            ["农业主要指标", None, None],
            ["指标", "单位", 2021.0],
            ["乡村户数", "万户", 1835.5],
            [None, None, None],
            ["行政区划", None, None],
            ["市别", "县", "乡"],
            ["全省", 76, 1120],
        ]
    )

    loaded = ExcelLoader()._normalize_sheet("Sheet1", dataframe)

    assert list(loaded.dataframe.columns) == ["指标", "单位", "2021"]
    assert loaded.dataframe.to_dict(orient="records") == [
        {"指标": "乡村户数", "单位": "万户", "2021": 1835.5}
    ]


def test_xlsx_loader_collects_merged_and_styled_structure_signals(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "农业主要指标"
    sheet["A2"] = "地区"
    sheet["B2"] = 2020
    sheet["C2"] = 2021
    sheet["A3"] = "广州"
    sheet["B3"] = 12
    sheet["C3"] = 14
    for cell in sheet[2]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font

    path = tmp_path / "styled.xlsx"
    workbook.save(path)

    loaded = ExcelLoader().load(str(path))[0]

    assert loaded.title == "农业主要指标"
    assert list(loaded.dataframe.columns) == ["地区", "2020", "2021"]
    assert loaded.dataframe.iloc[0].to_dict() == {"地区": "广州", "2020": 12, "2021": 14}


def test_xlsx_loader_splits_multiple_detected_regions_into_logical_sheets(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "农业主要指标"
    sheet["A2"] = "指标"
    sheet["B2"] = "单位"
    sheet["C2"] = 2021
    sheet["A3"] = "乡村户数"
    sheet["B3"] = "万户"
    sheet["C3"] = 1835.5
    sheet["A5"] = "行政区划"
    sheet["A6"] = "市别"
    sheet["B6"] = "县"
    sheet["C6"] = "乡"
    sheet["A7"] = "全省"
    sheet["B7"] = 76
    sheet["C7"] = 1120

    path = tmp_path / "multi-table.xlsx"
    workbook.save(path)

    loaded_sheets = ExcelLoader().load(str(path))

    assert [sheet.name for sheet in loaded_sheets] == ["Sheet1", "Sheet1 #2"]
    assert [sheet.title for sheet in loaded_sheets] == ["农业主要指标", "行政区划"]
    assert [list(sheet.dataframe.columns) for sheet in loaded_sheets] == [
        ["指标", "单位", "2021"],
        ["市别", "县", "乡"],
    ]
    assert loaded_sheets[0].dataframe.to_dict(orient="records") == [
        {"指标": "乡村户数", "单位": "万户", "2021": 1835.5}
    ]
    assert loaded_sheets[1].dataframe.to_dict(orient="records") == [
        {"市别": "全省", "县": 76, "乡": 1120}
    ]


def test_loader_skips_blank_sheets_before_importable_sheet(tmp_path) -> None:
    workbook = Workbook()
    blank_sheet = workbook.active
    blank_sheet.title = "Blank1"
    workbook.create_sheet("Blank2")
    data_sheet = workbook.create_sheet("Data")
    data_sheet["A1"] = "metric"
    data_sheet["B1"] = "value"
    data_sheet["A2"] = "street"
    data_sheet["B2"] = 8

    path = tmp_path / "blank-sheets.xlsx"
    workbook.save(path)
    workbook.close()

    loaded_sheets = ExcelLoader().load(str(path))

    assert [sheet.name for sheet in loaded_sheets] == ["Data"]
    assert loaded_sheets[0].dataframe.to_dict(orient="records") == [
        {"metric": "street", "value": 8}
    ]


def test_loader_keeps_complex_year_header_across_print_spacer_row() -> None:
    dataframe = pd.DataFrame(
        [
            ["1-2 国民经济和社会发展总量与速度指标", None, None, None, None, None, None],
            [None, None, None, None, None, None, None],
            ["指 标", "Item", 1978.0, 1990.0, 2000.0, "指数(2024为以下各年)", None],
            [None, None, None, None, None, 1978.0, 1990.0],
            [None, None, None, None, None, None, None],
            ["人口与就业", "Population and Employment", None, None, None, None, None],
            ["人口 (万人)", "Population (10 000 persons)", None, None, None, None, None],
            ["年末户籍总人口", "Registered Population at the Year-end", 5064.15, 6246.32, 7498.54, 202.36, 164.06],
        ]
    )

    loaded_sheets = ExcelLoader()._normalize_sheet_regions("Sheet1", dataframe)

    assert len(loaded_sheets) == 1
    assert list(loaded_sheets[0].dataframe.columns) == [
        "指标",
        "item",
        "1978",
        "1990",
        "2000",
        "指数_2024为以下各年_1978",
        "指数_2024为以下各年_1990",
    ]
    assert loaded_sheets[0].dataframe.iloc[-1].to_dict() == {
        "指标": "年末户籍总人口",
        "item": "Registered Population at the Year-end",
        "1978": 5064.15,
        "1990": 6246.32,
        "2000": 7498.54,
        "指数_2024为以下各年_1978": 202.36,
        "指数_2024为以下各年_1990": 164.06,
    }


def test_loader_keeps_four_row_year_header_and_skips_spacer_before_city_rows() -> None:
    dataframe = pd.DataFrame(
        [
            ["1-8 各市法人和产业活动单位数", None, None, None, None, None],
            ["Number of Corporate Units and Industrial Establishments by City", None, None, None, None, None],
            [None, None, None, None, None, None],
            ["单位: 个", None, None, None, None, "(unit)"],
            ["市 别", "City", 2023.0, None, 2024.0, None],
            [None, None, "法人单位数", "产业单位数", "法人单位数", "产业单位数"],
            [None, None, "Corporate", "Industrial", "Corporate", "Industrial"],
            [None, None, "Units", "Establishments", "Units", "Establishments"],
            [None, None, None, None, None, None],
            ["全省", "Provincial Total", 4708726, 5012232, 4989249, 5290000],
            ["韶关", "Shaoguan", 50698, 55833, 53545, 58687],
        ]
    )

    loaded_sheets = ExcelLoader()._normalize_sheet_regions("2-6", dataframe)

    assert len(loaded_sheets) == 1
    assert loaded_sheets[0].unit == "个"
    assert list(loaded_sheets[0].dataframe.columns) == [
        "市别",
        "city",
        "2023_法人单位数_corporate_units",
        "2023_产业单位数_industrial_establishments",
        "2024_法人单位数_corporate_units",
        "2024_产业单位数_industrial_establishments",
    ]
    assert loaded_sheets[0].dataframe.to_dict(orient="records") == [
        {
            "市别": "全省",
            "city": "Provincial Total",
            "2023_法人单位数_corporate_units": 4708726,
            "2023_产业单位数_industrial_establishments": 5012232,
            "2024_法人单位数_corporate_units": 4989249,
            "2024_产业单位数_industrial_establishments": 5290000,
        },
        {
            "市别": "韶关",
            "city": "Shaoguan",
            "2023_法人单位数_corporate_units": 50698,
            "2023_产业单位数_industrial_establishments": 55833,
            "2024_法人单位数_corporate_units": 53545,
            "2024_产业单位数_industrial_establishments": 58687,
        },
    ]


def test_loader_keeps_text_only_category_rows_after_header_spacer() -> None:
    dataframe = pd.DataFrame(
        [
            ["Report", None, None, None, None],
            [None, None, None, None, None],
            ["Label", "Item", 2000.0, 2010.0, 2024.0],
            [None, None, None, None, None],
            ["Group A", "Group A label", None, None, None],
            ["Category", "Category label", None, None, None],
            ["Alpha", "Alpha label", 97.6, 96.5, 98.7],
            ["Wrapped", "Wrapped label first line", None, None, None],
            [None, "Wrapped label second line", None, None, None],
            ["Beta", "Beta label", 1.6, 1.5, 0.7],
        ]
    )

    loaded_sheets = ExcelLoader()._normalize_sheet_regions("Sheet1", dataframe)

    assert len(loaded_sheets) == 1
    assert list(loaded_sheets[0].dataframe.columns) == [
        "label",
        "item",
        "2000",
        "2010",
        "2024",
    ]
    records = loaded_sheets[0].dataframe.to_dict(orient="records")
    assert [(record["label"], record["item"]) for record in records] == [
        ("Group A", "Group A label"),
        ("Category", "Category label"),
        ("Alpha", "Alpha label"),
        ("Wrapped", "Wrapped label first line"),
        (None, "Wrapped label second line"),
        ("Beta", "Beta label"),
    ]
    assert pd.isna(records[0]["2000"])
    assert pd.isna(records[1]["2024"])
    assert records[2] == {
        "label": "Alpha",
        "item": "Alpha label",
        "2000": 97.6,
        "2010": 96.5,
        "2024": 98.7,
    }
    assert records[5] == {
        "label": "Beta",
        "item": "Beta label",
        "2000": 1.6,
        "2010": 1.5,
        "2024": 0.7,
    }


def test_loader_does_not_treat_unit_symbol_as_title_when_title_is_above_blank_rows() -> None:
    dataframe = pd.DataFrame(
        [
            [None, "Report 1 continued", None, None, None],
            [None, None, None, None, None],
            ["Unit: %", None, None, None, "(%)"],
            ["Label", "Item", 2000.0, 2010.0, 2024.0],
            [None, None, None, None, None],
            ["Group A", "Group A label", None, None, None],
            ["Alpha", "Alpha label", 97.6, 96.5, 98.7],
        ]
    )

    loaded_sheets = ExcelLoader()._normalize_sheet_regions("Sheet1", dataframe)

    assert len(loaded_sheets) == 1
    assert loaded_sheets[0].title == "Report 1 continued"
    assert loaded_sheets[0].unit == "%"
    assert list(loaded_sheets[0].dataframe.columns) == [
        "label",
        "item",
        "2000",
        "2010",
        "2024",
    ]
