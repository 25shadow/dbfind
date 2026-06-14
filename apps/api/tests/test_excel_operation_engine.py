from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.services.excel_operation_engine import (
    DataFrameOperationEngine,
    WorkbookDesign,
    WorkbookOperationEngine,
)


def test_dataframe_operation_engine_filters_selects_sorts_and_rounds_values():
    dataframe = pd.DataFrame(
        [
            {"市别": "云浮", "街道": 8.123, "备注": "保留"},
            {"市别": "韶关", "街道": 10.987, "备注": "保留"},
            {"市别": "广州", "街道": 11.111, "备注": "排除"},
        ]
    )
    operations = [
        {"type": "filter_in", "column": "市别", "values": ["韶关", "云浮"]},
        {"type": "select_columns", "columns": ["市别", "街道"]},
        {"type": "sort_values", "by": ["街道"], "ascending": False},
        {"type": "round", "columns": ["街道"], "decimals": 2},
    ]

    result = DataFrameOperationEngine().apply(dataframe, operations)

    assert result.to_dict("records") == [
        {"市别": "韶关", "街道": 10.99},
        {"市别": "云浮", "街道": 8.12},
    ]


def test_dataframe_operation_engine_evaluates_pandas_expression_for_derived_columns():
    dataframe = pd.DataFrame(
        [
            {"城市": "广州", "街道数量": 143, "乡镇数量": 11},
            {"城市": "佛山", "街道数量": 42, "乡镇数量": 9},
        ]
    )

    result = DataFrameOperationEngine().apply(
        dataframe,
        [
            {
                "type": "eval_expression",
                "expression": "`城乡差值` = `街道数量` - `乡镇数量`",
            }
        ],
    )

    assert result.to_dict("records") == [
        {"城市": "广州", "街道数量": 143, "乡镇数量": 11, "城乡差值": 132},
        {"城市": "佛山", "街道数量": 42, "乡镇数量": 9, "城乡差值": 33},
    ]


def test_dataframe_operation_engine_uses_pandas_cleaning_operations():
    dataframe = pd.DataFrame(
        [
            {"城市": "广州", "销售额": "10.5", "备注": None},
            {"城市": "广州", "销售额": "10.5", "备注": None},
            {"城市": "佛山", "销售额": "8.25", "备注": "保留"},
            {"城市": "韶关", "销售额": None, "备注": "删除"},
        ]
    )

    result = DataFrameOperationEngine().apply(
        dataframe,
        [
            {"type": "rename_columns", "columns": {"销售额": "金额"}},
            {"type": "drop_duplicates", "subset": ["城市", "金额"]},
            {"type": "fillna", "values": {"备注": "待补充"}},
            {"type": "dropna", "subset": ["金额"]},
            {"type": "astype", "columns": {"金额": "float64"}},
            {"type": "query", "expression": "`金额` >= 9"},
        ],
    )

    assert result.to_dict("records") == [
        {"城市": "广州", "金额": 10.5, "备注": "待补充"},
    ]


def test_dataframe_operation_engine_uses_pandas_groupby_aggregation():
    dataframe = pd.DataFrame(
        [
            {"城市": "广州", "类别": "A", "金额": 10},
            {"城市": "广州", "类别": "B", "金额": 12},
            {"城市": "佛山", "类别": "A", "金额": 8},
        ]
    )

    result = DataFrameOperationEngine().apply(
        dataframe,
        [
            {
                "type": "groupby_agg",
                "by": ["城市"],
                "aggregations": {"金额": "sum", "类别": "count"},
            },
            {"type": "sort_values", "by": "金额_sum", "ascending": False},
        ],
    )

    assert result.to_dict("records") == [
        {"城市": "广州", "金额_sum": 22, "类别_count": 2},
        {"城市": "佛山", "金额_sum": 8, "类别_count": 1},
    ]


def test_workbook_operation_engine_writes_designed_workbook(tmp_path):
    output_path = tmp_path / "agent-output.xlsx"
    dataframe = pd.DataFrame(
        [
            {"市别": "韶关", "街道": 10.99},
            {"市别": "云浮", "街道": 8.12},
        ]
    )

    WorkbookOperationEngine().write_workbook(
        output_path,
        [("对比表", dataframe)],
        WorkbookDesign(
            freeze_header=True,
            autofilter=True,
            header_fill="E8F1FF",
            number_formats={"街道": "0.00"},
        ),
    )

    workbook = load_workbook(output_path)
    sheet = workbook["对比表"]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:B3"
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].fill.fgColor.rgb == "FFE8F1FF"
    assert sheet["B2"].number_format == "0.00"
    assert sheet["B3"].value == 8.12


def test_workbook_operation_engine_uses_xlsxwriter_tables_and_conditional_formats(tmp_path):
    output_path = tmp_path / "agent-report.xlsx"
    dataframe = pd.DataFrame(
        [
            {"城市": "广州", "街道数量": 143, "差值": 132},
            {"城市": "佛山", "街道数量": 11, "差值": 132},
        ]
    )

    WorkbookOperationEngine().write_workbook(
        output_path,
        [("计算表", dataframe)],
        WorkbookDesign(
            as_table=True,
            table_style="Table Style Medium 2",
            number_formats={"街道数量": "0", "差值": "0"},
            conditional_formats=[
                {
                    "column": "差值",
                    "type": "cell",
                    "criteria": ">",
                    "value": 0,
                    "format": {"bg_color": "#E2F0D9", "font_color": "#1F7A3A"},
                }
            ],
        ),
    )

    workbook = load_workbook(output_path)
    sheet = workbook["计算表"]
    assert len(sheet.tables) == 1
    assert next(iter(sheet.tables.values())).ref == "A1:C3"
    assert sheet["B2"].number_format == "0"
    assert sheet["C2"].number_format == "0"
    assert len(sheet.conditional_formatting) == 1


def test_workbook_operation_engine_uses_xlsxwriter_charts(tmp_path):
    output_path = tmp_path / "agent-chart.xlsx"
    dataframe = pd.DataFrame(
        [
            {"城市": "广州", "金额": 10},
            {"城市": "佛山", "金额": 8},
        ]
    )

    WorkbookOperationEngine().write_workbook(
        output_path,
        [("销售", dataframe)],
        WorkbookDesign(
            charts=[
                {
                    "type": "column",
                    "title": "城市销售额",
                    "categoriesColumn": "城市",
                    "valuesColumn": "金额",
                    "position": "D2",
                }
            ],
        ),
    )

    workbook = load_workbook(output_path)
    sheet = workbook["销售"]
    assert len(sheet._charts) == 1
