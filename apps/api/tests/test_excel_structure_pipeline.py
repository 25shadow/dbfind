from openpyxl import Workbook

from app.schemas.table_structure import TableStructurePlan
from app.services.excel_structure_pipeline import ExcelStructurePipeline


class FakeVisionPlanner:
    def __init__(self, plan):
        self._plan = plan
        self.calls = []

    def availability_issue(self):
        return None

    def plan(self, grid, rendering):
        self.calls.append((grid, rendering))
        return self._plan


def test_structure_pipeline_extracts_table_from_vlm_coordinate_plan(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Complex"
    sheet["A1"] = "Report Title"
    sheet["A2"] = "Report Subtitle"
    sheet["C5"] = 2023
    sheet["E5"] = 2024
    sheet["C6"] = "Measure A"
    sheet["D6"] = "Measure B"
    sheet["E6"] = "Measure A"
    sheet["F6"] = "Measure B"
    sheet["C7"] = "Count"
    sheet["D7"] = "Ratio"
    sheet["E7"] = "Count"
    sheet["F7"] = "Ratio"
    sheet["A9"] = "Alpha"
    sheet["B9"] = "Alpha label"
    sheet["C9"] = 100
    sheet["D9"] = 0.2
    sheet["E9"] = 120
    sheet["F9"] = 0.3
    sheet["A10"] = "Beta"
    sheet["B10"] = "Beta label"
    sheet["C10"] = 210
    sheet["D10"] = 0.4
    sheet["E10"] = 230
    sheet["F10"] = 0.5

    path = tmp_path / "complex.xlsx"
    workbook.save(path)

    planner = FakeVisionPlanner(
        TableStructurePlan(
            tableRegion="A5:F10",
            titleRows=[1, 2],
            unitCells=[],
            headerRows=[5, 6, 7],
            dataStartRow=9,
            dataEndRow=10,
            rowHeaderColumns=["A", "B"],
            valueColumns=["C", "D", "E", "F"],
            categoryRows=[],
            orientation="wide_year_table",
            confidence=0.91,
            source="vlm",
        )
    )

    results = ExcelStructurePipeline(vision_planner=planner).parse(str(path))

    assert len(results) == 1
    result = results[0]
    assert result.status == "ready"
    assert result.sheet_name == "Complex"
    assert result.block_region == "A5:F10"
    assert result.plan is not None
    assert result.plan.source == "vlm"
    assert result.plan.table_region == "A5:F10"
    assert result.title == "Report Title"
    assert result.subtitle == "Report Subtitle"
    assert result.quality.confidence == "medium"
    assert list(result.dataframe.columns) == [
        "A",
        "B",
        "2023_Measure_A_Count",
        "2023_Measure_B_Ratio",
        "2024_Measure_A_Count",
        "2024_Measure_B_Ratio",
    ]
    assert result.dataframe.iloc[1].to_dict()["2024_Measure_B_Ratio"] == 0.5
    assert planner.calls


def test_structure_pipeline_preserves_non_data_raw_content_blocks(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notes"
    sheet["A1"] = "Title"
    sheet["A2"] = "Name"
    sheet["B2"] = 2024
    sheet["A3"] = "Alpha"
    sheet["B3"] = 12
    sheet["A5"] = "注：本行是原文备注。"
    sheet["A6"] = "Note: original note text."
    path = tmp_path / "notes.xlsx"
    workbook.save(path)
    planner = FakeVisionPlanner(
        TableStructurePlan(
            tableRegion="A1:B3",
            titleRows=[1],
            unitCells=[],
            headerRows=[2],
            dataStartRow=3,
            dataEndRow=3,
            rowHeaderColumns=["A"],
            valueColumns=["B"],
            categoryRows=[],
            orientation="wide_table",
            confidence=0.9,
            source="vlm",
        )
    )

    result = ExcelStructurePipeline(vision_planner=planner).parse(str(path))[0]

    assert result.status == "ready"
    assert [block["text"] for block in result.raw_content_blocks] == [
        "Title",
        "Name 2024",
        "注：本行是原文备注。",
        "Note: original note text.",
    ]
    assert result.raw_content_blocks[-2]["region"] == "A5:A5"


def test_structure_pipeline_result_keeps_all_extracted_rows(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rows"
    sheet["A1"] = "Name"
    sheet["B1"] = "Value"
    for index in range(1, 23):
        row = index + 1
        sheet.cell(row=row, column=1).value = f"City {index}"
        sheet.cell(row=row, column=2).value = index
    path = tmp_path / "rows.xlsx"
    workbook.save(path)
    planner = FakeVisionPlanner(
        TableStructurePlan(
            tableRegion="A1:B23",
            titleRows=[],
            unitCells=[],
            headerRows=[1],
            dataStartRow=2,
            dataEndRow=23,
            rowHeaderColumns=["A"],
            valueColumns=["B"],
            categoryRows=[],
            orientation="wide_table",
            confidence=0.9,
            source="vlm",
        )
    )

    result = ExcelStructurePipeline(vision_planner=planner).parse(str(path))[0]

    assert len(result.dataframe) == 22
    assert result.dataframe.iloc[-1].to_dict() == {"Name": "City 22", "Value": 22}


def test_structure_pipeline_skips_empty_sheets(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "Name"
    sheet["B1"] = "Value"
    sheet["A2"] = "Alpha"
    sheet["B2"] = 12
    workbook.create_sheet("Empty")
    path = tmp_path / "with-empty-sheet.xlsx"
    workbook.save(path)

    planner = FakeVisionPlanner(
        TableStructurePlan(
            tableRegion="A1:B2",
            titleRows=[],
            unitCells=[],
            headerRows=[1],
            dataStartRow=2,
            dataEndRow=2,
            rowHeaderColumns=["A"],
            valueColumns=["B"],
            categoryRows=[],
            orientation="wide_table",
            confidence=0.9,
            source="vlm",
        )
    )

    results = ExcelStructurePipeline(vision_planner=planner).parse(str(path))

    assert [result.sheet_name for result in results] == ["Data"]
    assert [call[0].sheet_name for call in planner.calls] == ["Data"]


def test_structure_pipeline_keeps_vlm_metadata_for_review_results(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review"
    sheet["A1"] = "2-8  分行业地区生产总值"
    sheet["A2"] = "Gross Domestic Product by Sector"
    sheet["A4"] = "单位：亿元"
    sheet["F4"] = "(100 million yuan)"
    sheet["A5"] = "行业"
    sheet["B5"] = "Sector"
    sheet["A7"] = "地区生产总值"
    sheet["B7"] = "Gross Domestic Product"
    path = tmp_path / "review.xlsx"
    workbook.save(path)

    planner = FakeVisionPlanner(
        TableStructurePlan(
            tableRegion="A1:F7",
            titleRows=[1, 2],
            unitCells=["F4"],
            headerRows=[],
            dataStartRow=7,
            dataEndRow=7,
            rowHeaderColumns=["A", "B"],
            valueColumns=[],
            categoryRows=[],
            orientation="wide_table",
            confidence=0.9,
            source="vlm",
        )
    )

    result = ExcelStructurePipeline(vision_planner=planner).parse(str(path))[0]

    assert result.status == "needs_review"
    assert result.title == "2-8 分行业地区生产总值"
    assert result.subtitle == "Gross Domestic Product by Sector"
    assert result.unit == "单位：亿元 (100 million yuan)"


def test_structure_pipeline_marks_unplanned_blocks_for_review(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notes"
    sheet["A1"] = "Only notes"
    sheet["A2"] = "No numeric data"

    path = tmp_path / "notes.xlsx"
    workbook.save(path)

    results = ExcelStructurePipeline(vision_planner=FakeVisionPlanner(None)).parse(str(path))

    assert len(results) == 1
    assert results[0].status == "needs_review"
    assert results[0].plan is None
    assert "vlm_structure_plan_not_found" in results[0].issues


def test_structure_pipeline_does_not_fallback_to_old_rules_when_vlm_unavailable(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "WouldPassRules"
    sheet["C5"] = 2023
    sheet["C6"] = "Measure"
    sheet["A9"] = "Alpha"
    sheet["B9"] = "Alpha label"
    sheet["C9"] = 100
    sheet["A10"] = "Beta"
    sheet["B10"] = "Beta label"
    sheet["C10"] = 210

    path = tmp_path / "would-pass-rules.xlsx"
    workbook.save(path)

    results = ExcelStructurePipeline(vision_planner=FakeVisionPlanner(None)).parse(str(path))

    assert len(results) == 1
    assert results[0].status == "needs_review"
    assert results[0].plan is None
    assert "vlm_structure_plan_not_found" in results[0].issues
